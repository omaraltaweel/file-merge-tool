import streamlit as st
import pandas as pd
import os
import time
import sys
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.comments import Comment
from config import TEMPLATE_EXPECTED_HEADERS, UNWANTED_COLUMNS

# === Password Protection ===
# PASSWORD = "0000"

# if "authenticated" not in st.session_state:
    # st.session_state.authenticated = False

# if not st.session_state.authenticated:
    # pw = st.text_input("🔒 Enter password to access the tool", type="password")
    # if pw == PASSWORD:
        # st.session_state.authenticated = True
        # st.success("✅ Access granted. Please wait...")
        # st.rerun()
    # elif pw:
        # st.error("❌ Incorrect password.")
        # st.stop()
    # else:
        # st.stop()
        
def wait_until_file_ready(filepath, timeout=10):
    start_time = time.time()
    last_size = -1
    while time.time() - start_time < timeout:
        if os.path.exists(filepath):
            current_size = os.path.getsize(filepath)
            if current_size == last_size:
                return True
            last_size = current_size
        time.sleep(0.5)
    return False

st.set_page_config(page_title="Excel File Merge Tool", layout="wide")
st.title("📎 Standard Materials Import Tool")

if "saved_path" not in st.session_state:
    st.session_state.saved_path = None
if "merge_stats" not in st.session_state:
    st.session_state.merge_stats = {}
if "uploaded_filenames" not in st.session_state:
    st.session_state.uploaded_filenames = []

uploaded_files = st.file_uploader(
    "Upload Excel files (must contain a sheet named 'Standard Materials')",
    type=["xlsx", "xlsm"],
    accept_multiple_files=True
)

current_filenames = [f.name for f in uploaded_files] if uploaded_files else []
changed_files = current_filenames != st.session_state.uploaded_filenames

# Clear messages and state if no files are uploaded
if not uploaded_files:
    st.session_state.saved_path = None
    st.session_state.merge_stats = {}
    st.session_state.uploaded_filenames = []

if st.button("📂 Generate and Open Merged Excel File"):
    if uploaded_files and changed_files:
        all_data = []
        validation_errors = {}

        for file in uploaded_files:
            try:
                xl = pd.ExcelFile(file)
                sheet_name = next((s for s in xl.sheet_names if s.lower() == "standard materials"), None)
                if not sheet_name:
                    validation_errors[file.name] = ["Missing 'Standard Materials' sheet"]
                    continue

                df = xl.parse(sheet_name, dtype=str).fillna("")
                original_headers = [str(col).strip() for col in df.columns]
                lower_headers = [col.lower() for col in original_headers]

                errors = []
                for i, expected in enumerate(TEMPLATE_EXPECTED_HEADERS):
                    if i >= len(lower_headers):
                        errors.append(f"Missing column {i+1}: '{expected}'")
                    elif lower_headers[i] != expected.lower():
                        errors.append(f"Column {i+1}: Found '{original_headers[i]}', expected '{expected}'")

                if errors:
                    validation_errors[file.name] = errors
                    continue

                df.insert(0, "S.I.", "")
                df["SourceFile"] = file.name
                df = df.map(lambda x: x.strip() if isinstance(x, str) else x)
                all_data.append((file.name, df))

            except Exception as e:
                validation_errors[file.name] = [str(e)]

        if validation_errors:
            st.error("❌ Some files failed validation:")
            for fname, issues in validation_errors.items():
                st.markdown(f"**{fname}**")
                for issue in issues:
                    st.markdown(f"- {issue}")
            st.stop()

        if not all_data:
            st.warning("No valid data available to merge.")
            st.stop()

        canonical_headers = {}
        for _, df in all_data:
            for col in df.columns:
                key = col.strip().lower()
                if key not in canonical_headers:
                    canonical_headers[key] = col.strip()

        for idx, (fname, df) in enumerate(all_data):
            df.columns = [canonical_headers.get(col.strip().lower(), col) for col in df.columns]
            all_data[idx] = (fname, df)

        combined_df = pd.concat([df for _, df in all_data], ignore_index=True)
        combined_df.sort_values(by="Material_ID", inplace=True, ignore_index=True)
        duplicate_ids = combined_df[combined_df.duplicated("Material_ID", keep=False)]

        now = datetime.now()
        username = os.getenv("USERNAME", "User")
        timestamp = now.strftime("%Y%m%d_%H%M%S")
        output_filename = f"MergedOutput_{username}__{timestamp}.xlsx"

        if getattr(sys, 'frozen', False):
            base_dir = os.path.dirname(sys.executable)
        else:
            base_dir = os.path.dirname(os.path.abspath(__file__))

        output_dir = os.path.join(base_dir, "Output")
        os.makedirs(output_dir, exist_ok=True)

        saved_path = os.path.join(output_dir, output_filename)
        combined_df.to_excel(saved_path, index=False, sheet_name="Merged Data")

        wb = load_workbook(saved_path)
        ws = wb.active

        for row_idx, cell in enumerate(ws["B"][1:], start=2):
            try:
                source_file = combined_df.at[row_idx - 2, "SourceFile"]
                cell.comment = Comment(f"SourceFile: {source_file}", "MergeTool")
            except Exception:
                continue

        for col_idx, cell in enumerate(ws[1], start=1):
            if cell.value == "SourceFile":
                ws.delete_cols(col_idx)
                break

        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=False)
                cell.fill = PatternFill(fill_type=None)

        material_ids = [cell.value for cell in ws["B"][1:] if cell.value]
        duplicates = {x for x in material_ids if material_ids.count(x) > 1}
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        for cell in ws["B"][1:]:
            if cell.value in duplicates:
                cell.fill = red_fill

        non_empty_highlight_cols = []
        deleted_cols = []
        header_row = [cell.value for cell in ws[1]]

        for col_idx, header in enumerate(header_row, start=1):
            if header and header.strip().lower() in UNWANTED_COLUMNS:
                col_data = list(ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2, values_only=True))[0]
                if all(v in ("", None) for v in col_data):
                    deleted_cols.append(col_idx)
                else:
                    non_empty_highlight_cols.append(header)

        for col_idx in sorted(deleted_cols, reverse=True):
            ws.delete_cols(col_idx)

        for col_idx, header in enumerate([cell.value for cell in ws[1]], start=1):
            if header and header.strip().lower() in UNWANTED_COLUMNS:
                col_data = list(ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2, values_only=True))[0]
                if any(v not in ("", None) for v in col_data):
                    ws.cell(row=1, column=col_idx).fill = PatternFill(start_color="FF6666", end_color="FF6666", fill_type="solid")

        max_row, max_col = ws.max_row, ws.max_column
        if max_row >= 2 and max_col >= 1:
            table_range = f"A1:{get_column_letter(max_col)}{max_row}"
            table = Table(displayName="MergedDataTable", ref=table_range)
            style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
            table.tableStyleInfo = style
            ws.add_table(table)

        ws.freeze_panes = "A2"

        for col in ws.columns:
            max_len = max((len(str(cell.value)) for cell in col if cell.value), default=0)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 50)

        wb.save(saved_path)
        st.session_state.saved_path = saved_path
        st.session_state.merge_stats = {
            "files_merged": len(all_data),
            "rows_merged": len(combined_df),
            "duplicates": not duplicate_ids.empty,
            "non_empty_columns": non_empty_highlight_cols,
            "output_path": saved_path
        }
        st.session_state.uploaded_filenames = current_filenames

        if wait_until_file_ready(saved_path):
            os.startfile(saved_path)
        else:
            st.error("⚠️ File was not ready to open. Please try again.")

    elif not uploaded_files:
        st.warning("⚠️ No files uploaded. Please upload files before generating.")

if st.session_state.saved_path and st.session_state.merge_stats:
    stats = st.session_state.merge_stats
    st.success(f"✅ Files merged: {stats['files_merged']}")
    st.success(f"✅ Rows merged: {stats['rows_merged']}")
    st.success(f"📁 Merged file saved to: `{os.path.abspath(stats['output_path'])}`")

    if stats['duplicates']:
        st.warning("⚠️ Duplicate Material_IDs were found and highlighted in RED.")
    if stats['non_empty_columns']:
        st.warning("⚠️ Some UNWANTED_COLUMNS contain data and could not be removed:\n- " + ", ".join(stats['non_empty_columns']))
